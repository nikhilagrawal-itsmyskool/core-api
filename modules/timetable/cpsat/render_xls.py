#!/usr/bin/env python3
"""
Render a CP-SAT timetable run to a CONSOLIDATED master Excel workbook — one sheet,
all class-sections stacked, mirroring the school's hand-kept timetable layout.

Layout (option A — consolidated):
  - rows   = classes (with sections), one row-block each
  - columns= periods across the full day; teaching slots are numbered I, II, III...
             and fixed slots (Assembly / Reserved / Lunch / Activity) are kept in place
  - cell   = day-consolidated: identical days collapse into one entry with a day-set
             annotation, e.g. "Science (1-6)" / "Sanskrit (3,4)"; day-varying periods
             stack each group (e.g. a Saturday "Activity (6)" beside "Maths (1-5)").
             Teacher shown by full name. Days are 1=Mon .. 6=Sat.

Reads a run folder (cpsat/<run-id>/): solution.json, solver-input.json, registration.json.

Usage:
    python3 render_xls.py --run-dir cpsat/<run-id>          # writes <run-dir>/timetable.xlsx
    python3 render_xls.py --in solver-input.json --solution solution.json \
                          --registration registration.json --out timetable.xlsx

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import json
import os
import re
from collections import OrderedDict, defaultdict

DAY_NAMES = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat", 7: "Sun"}
ROMAN_SEQ = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII",
             "XIII", "XIV", "XV"]
FIXED = {"break": "Break", "lunch": "Lunch", "assembly": "Assembly",
         "reserved": "Reserved", "activity": "Activity", "registration": "Reg"}

_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7,
          "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}


def _class_sort_key(label):
    """Sort 'XI-A' after 'X-A', not lexically between 'IV' and 'V'."""
    m = re.match(r"^\s*(XII|XI|IX|VIII|VII|VI|IV|III|II|X|V|I)", label or "")
    grade = _ROMAN.get(m.group(1), 0) if m else 0
    rest = label[m.end():] if m else (label or "")
    return (grade, rest)


def classes_of(p):
    ids = p.get("classIds") or []
    return ids if len(ids) > 0 else [p["classId"]]


def subj_name(sub_l, sid):
    """'Mathematics (041-MATHS)' -> 'Mathematics' (drop the code for a clean cell)."""
    lab = sub_l.get(sid) or sid[:6]
    return re.sub(r"\s*\([^)]*\)\s*$", "", lab)


def teacher_name(tch_l, tid):
    return (tch_l.get(tid) or tid[:4]) if tid else ""


def fmt_days(ds):
    """[1,2,3,4,5,6] -> '1-6' ; [3,4] -> '3,4' ; [1,2,5,6] -> '1-2,5-6'."""
    ds = sorted(set(ds))
    parts, i = [], 0
    while i < len(ds):
        j = i
        while j + 1 < len(ds) and ds[j + 1] == ds[j] + 1:
            j += 1
        parts.append(f"{ds[i]}-{ds[j]}" if j > i else f"{ds[i]}")
        i = j + 1
    return ",".join(parts)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", dest="run_dir", help="folder with solver-input/solution/registration json")
    ap.add_argument("--in", dest="inp", help="solver-input.json (default <run-dir>/solver-input.json)")
    ap.add_argument("--solution", dest="sol", help="solution.json (default <run-dir>/solution.json)")
    ap.add_argument("--registration", dest="reg", help="registration.json (default <run-dir>/registration.json)")
    ap.add_argument("--out", dest="out", help="output xlsx (default <run-dir>/timetable.xlsx)")
    args = ap.parse_args()

    rd = args.run_dir
    inp = args.inp or (os.path.join(rd, "solver-input.json") if rd else None)
    sol = args.sol or (os.path.join(rd, "solution.json") if rd else None)
    reg = args.reg or (os.path.join(rd, "registration.json") if rd else None)
    out = args.out or (os.path.join(rd, "timetable.xlsx") if rd else "timetable.xlsx")
    if not inp or not sol:
        ap.error("need --run-dir, or both --in and --solution")

    with open(inp) as f:
        payload = json.load(f)
    data = payload.get("input", payload)
    labels = payload.get("labels", {})
    meta = payload.get("meta", {})
    cls_l = labels.get("class", {})
    sub_l = labels.get("subject", {})
    tch_l = labels.get("teacher", {})

    with open(sol) as f:
        placements = json.load(f)
    registration = []
    if reg and os.path.exists(reg):
        with open(reg) as f:
            registration = json.load(f)

    # ---- grid: (day,seq)->slotType, slotId->(day,seq), ordered seqs/days ----------
    grid = data["grid"]
    days = sorted(d["dayOfWeek"] for d in grid["days"])
    slot_type, slot_loc = {}, {}
    seqset = set()
    for d in grid["days"]:
        dow = d["dayOfWeek"]
        for s in d["slots"]:
            slot_type[(dow, s["sequence"])] = s["slotType"]
            slot_loc[s["slotId"]] = (dow, s["sequence"])
            seqset.add(s["sequence"])
    all_seqs = sorted(seqset)

    # ---- per-(class,day,seq) content ----------------------------------------------
    by_class_day = defaultdict(list)   # (c,d,seq) -> [(subject, teacher)]
    reg_by_class_day = {}              # (c,d,seq) -> teacher (registration)
    for p in placements:
        d = p["dayOfWeek"]
        for k in range(p["size"]):
            seq = p["startSequence"] + k
            for c in classes_of(p):
                for o in p["offerings"]:
                    by_class_day[(c, d, seq)].append(
                        (subj_name(sub_l, o["subjectId"]), teacher_name(tch_l, o.get("teacherId"))))
    for e in registration:
        loc = slot_loc.get(e["timeSlotId"])
        if loc:
            reg_by_class_day[(e["classId"], loc[0], loc[1])] = teacher_name(tch_l, e["teacherId"])

    # ---- column model: teaching seqs -> I, II, ... ; others -> their type label ----
    seq_label, teaching_cols = {}, set()
    ti = 0
    for seq in all_seqs:
        if any(slot_type.get((d, seq)) == "teaching" for d in days):
            seq_label[seq] = ROMAN_SEQ[ti] if ti < len(ROMAN_SEQ) else f"P{seq}"
            teaching_cols.add(seq)
            ti += 1
        else:
            types = [slot_type.get((d, seq)) for d in days if slot_type.get((d, seq))]
            seq_label[seq] = FIXED.get(types[0], types[0].title()) if types else f"P{seq}"

    # ---- consolidate one (class, seq) cell across days ----------------------------
    def cell_for(c, seq):
        col_days = [d for d in days if slot_type.get((d, seq)) is not None]
        groups = OrderedDict()  # key -> {"days":[], "disp":(subjects, teachers), "kind":..}
        for d in col_days:
            st = slot_type.get((d, seq))
            if st == "teaching":
                offs = by_class_day.get((c, d, seq))
                if not offs:
                    continue  # free teaching period
                subs = "/".join(s for s, _ in offs)
                tchs = "/".join(t for _, t in offs if t)
                key, disp, kind = ("teach", subs, tchs), (subs, tchs), "teach"
            elif (c, d, seq) in reg_by_class_day:
                t = reg_by_class_day[(c, d, seq)]
                key, disp, kind = ("reg", t), ("Reg", t), "fixed"
            else:
                lab = FIXED.get(st, st.title())
                key, disp, kind = ("fixed", lab), (lab, ""), "fixed"
            g = groups.setdefault(key, {"days": [], "disp": disp, "kind": kind})
            g["days"].append(d)

        lines = []
        for g in groups.values():
            subs, tchs = g["disp"]
            # fixed slot spanning the whole column needs no day annotation (e.g. "Lunch");
            # teaching always annotates, and partial fixed (e.g. Saturday Activity) too.
            annotate = not (g["kind"] == "fixed" and len(g["days"]) == len(col_days))
            head = f"{subs} ({fmt_days(g['days'])})" if annotate else subs
            lines.append(head + (f"\n{tchs}" if tchs else ""))
        return "\n".join(lines)

    # ---- render workbook ----------------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Master"
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    fixed_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)

    header = ["Class"] + [seq_label[seq] for seq in all_seqs]
    ws.append(header)
    for ci in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    class_ids = sorted(cls_l.keys(), key=lambda c: _class_sort_key(cls_l.get(c, c))) \
        or data.get("classIds", [])

    r = 2
    for c in class_ids:
        ws.cell(row=r, column=1, value=cls_l.get(c, c)).alignment = center
        ws.cell(row=r, column=1).font = Font(bold=True, size=9)
        ws.cell(row=r, column=1).border = border
        maxlines = 1
        for ci, seq in enumerate(all_seqs, start=2):
            txt = cell_for(c, seq)
            cell = ws.cell(row=r, column=ci, value=txt)
            cell.alignment = top
            cell.border = border
            cell.font = Font(size=8)
            if seq not in teaching_cols:
                cell.fill = fixed_fill
            maxlines = max(maxlines, txt.count("\n") + 1)
        ws.row_dimensions[r].height = max(16, maxlines * 11 + 4)
        r += 1

    ws.column_dimensions["A"].width = 16
    for ci, seq in enumerate(all_seqs, start=2):
        ws.column_dimensions[get_column_letter(ci)].width = 20 if seq in teaching_cols else 10
    ws.freeze_panes = "B2"

    wb.save(out)
    title = meta.get("runId", "")
    print(f"wrote {out}{f' (run {title})' if title else ''}: "
          f"{len(class_ids)} classes x {len(all_seqs)} period columns")


if __name__ == "__main__":
    main()
