#!/usr/bin/env python3
"""
Render a CP-SAT run to the SCHOOL'S hand-made block style — a drop-in alternative to
render_xls.py that mirrors `Time Table 2026-27.xlsx` (the format the school actually uses).

Layout (per the sampled workbook):
  - one **12-row block per class**; the class name is centred in column 1 of its block.
  - period columns are the teaching periods numbered I, II, III … with a red **Break** column
    at the lunch split and a trailing empty **Diary** column. Assembly / reserved slots are
    not shown (matching the hand file).
  - each (class, period) cell **stacks** its day-varying entries down the block rows:
    `Subject (day-set)` on one row, the **teacher's full name** on the next
    (e.g. `Physics (1-6)` / `Mr. Prakhar Mishra`). Saturday `activity` shows as `Activity (6)`.
  - colours reproduce the source exactly: each class block is one flat theme colour (the
    palette cycles themes 9→2 across tints .80/.60/.40), the Break column is solid red, and the
    header is a banded colour. Font is Times New Roman throughout. No clock-time row.

Reads a run folder (cpsat/<run-id>/): solution.json, solver-input.json (registration.json is
not shown in this style). Same CLI as render_xls.py.

Usage:
    python3 render_xls_school.py --run-dir cpsat/<run-id>          # -> <run-dir>/timetable.xlsx
    python3 render_xls_school.py --in solver-input.json --solution solution.json --out out.xlsx

Requires: openpyxl
"""
import argparse
import json
import os
import re
from collections import OrderedDict, defaultdict

DAY_NAMES = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat", 7: "Sun"}
ROMAN_SEQ = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII",
             "XIII", "XIV", "XV"]
_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7,
          "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}

BLOCK_ROWS = 12          # rows per class block (matches the source)

# Per-class block fill, sampled from Time Table 2026-27.xlsx: (theme, tint), cycled by block.
BLOCK_PALETTE = [
    (9, 0.80), (8, 0.80), (7, 0.80), (6, 0.80), (5, 0.80), (4, 0.80), (3, 0.80), (2, -0.10),
    (9, 0.60), (8, 0.60), (7, 0.60), (6, 0.60), (5, 0.60), (4, 0.60), (3, 0.60), (2, -0.25),
    (9, 0.40), (8, 0.40), (7, 0.40), (6, 0.40),
]
HEADER_THEME, HEADER_TINT = 7, 0.60
RED_ARGB = "FFFF0000"
FONT_NAME = "Times New Roman"


def _class_sort_key(label):
    """Order like the hand file: senior grades first (XI, X, IX … I), sections A→B within
    a grade. So a block's palette colour also lines up with the source (XI = theme9/.80)."""
    m = re.match(r"^\s*(XII|XI|IX|VIII|VII|VI|IV|III|II|X|V|I)", label or "")
    grade = _ROMAN.get(m.group(1), 0) if m else 0
    rest = label[m.end():] if m else (label or "")
    return (-grade, rest)


def classes_of(p):
    ids = p.get("classIds") or []
    return ids if len(ids) > 0 else [p["classId"]]


def subj_name(sub_l, sid):
    """'Mathematics (041-MATHS)' -> 'Mathematics' (drop the code for a clean cell)."""
    lab = sub_l.get(sid) or sid[:6]
    return re.sub(r"\s*\([^)]*\)\s*$", "", lab)


def teacher_name(tch_l, tid):
    return (tch_l.get(tid) or tid[:4]) if tid else ""


def fmt_days(ds):
    """[1,2,3,4,5,6] -> '1-6' ; [3,4] -> '3,4' ; [1,2,5,6] -> '1-2,5-6'."""
    ds = sorted(set(ds))
    parts, i = [], 0
    while i < len(ds):
        j = i
        while j + 1 < len(ds) and ds[j + 1] == ds[j] + 1:
            j += 1
        parts.append(f"{ds[i]}-{ds[j]}" if j > i else f"{ds[i]}")
        i = j + 1
    return ",".join(parts)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", dest="run_dir", help="folder with solver-input/solution json")
    ap.add_argument("--in", dest="inp", help="solver-input.json (default <run-dir>/solver-input.json)")
    ap.add_argument("--solution", dest="sol", help="solution.json (default <run-dir>/solution.json)")
    ap.add_argument("--out", dest="out", help="output xlsx (default <run-dir>/timetable.xlsx)")
    args = ap.parse_args()

    rd = args.run_dir
    inp = args.inp or (os.path.join(rd, "solver-input.json") if rd else None)
    sol = args.sol or (os.path.join(rd, "solution.json") if rd else None)
    out = args.out or (os.path.join(rd, "timetable.xlsx") if rd else "timetable.xlsx")
    if not inp or not sol:
        ap.error("need --run-dir, or both --in and --solution")

    with open(inp) as f:
        payload = json.load(f)
    data = payload.get("input", payload)
    labels = payload.get("labels", {})
    meta = payload.get("meta", {})
    cls_l = labels.get("class", {})
    sub_l = labels.get("subject", {})
    tch_l = labels.get("teacher", {})
    with open(sol) as f:
        placements = json.load(f)

    # ---- grid: (day,seq)->slotType, ordered seqs/days -----------------------------
    grid = data["grid"]
    days = sorted(d["dayOfWeek"] for d in grid["days"])
    slot_type = {}
    seqset = set()
    for d in grid["days"]:
        dow = d["dayOfWeek"]
        for s in d["slots"]:
            slot_type[(dow, s["sequence"])] = s["slotType"]
            seqset.add(s["sequence"])
    all_seqs = sorted(seqset)

    # ---- per-(class,day,seq) taught offerings -------------------------------------
    by_class_day = defaultdict(list)   # (c,d,seq) -> [(subject, teacher)]
    for p in placements:
        d = p["dayOfWeek"]
        for k in range(p["size"]):
            seq = p["startSequence"] + k
            for c in classes_of(p):
                for o in p["offerings"]:
                    by_class_day[(c, d, seq)].append(
                        (subj_name(sub_l, o["subjectId"]), teacher_name(tch_l, o.get("teacherId"))))

    # ---- column plan: teaching seq -> Roman column; lunch -> Break; else skipped ---
    # returns list of (label, kind) and a map teaching-seq -> 1-based output column.
    plan = [("Class", "class")]
    seq_col = {}
    ti = 0
    for seq in all_seqs:
        types = {slot_type.get((d, seq)) for d in days if slot_type.get((d, seq))}
        if "teaching" in types:
            plan.append((ROMAN_SEQ[ti] if ti < len(ROMAN_SEQ) else f"P{seq}", "teach"))
            seq_col[seq] = len(plan)          # 1-based column index
            ti += 1
        elif "lunch" in types or "break" in types:
            plan.append(("Break", "break"))
    plan.append(("Diary", "diary"))
    ncols = len(plan)
    break_cols = [i + 1 for i, (_, kind) in enumerate(plan) if kind == "break"]

    def day_groups(c, seq):
        """Group a (class, teaching column) BY SUBJECT across ALL its days (contiguous or not),
        one merged cell per subject sized to its day-count. Rows do NOT map to physical days —
        groups are stacked in order of their earliest day. Returns an ordered list of
        (days_list, subject, teachers). Free/non-teaching days contribute nothing (leaving
        blank rows at the bottom). Teachers are the ordered-unique union across the subject's
        days (so a split lists each teacher once)."""
        groups = OrderedDict()   # subject -> {"days": [...], "teachers": [...]}
        for d in days:           # ascending -> first appearance == earliest day == stack order
            st = slot_type.get((d, seq))
            if st == "teaching":
                offs = by_class_day.get((c, d, seq))
                if not offs:
                    continue                       # free teaching period
                subs = "/".join(s for s, _ in offs)
                tch = [t for _, t in offs if t]
            elif st == "activity":
                subs, tch = "Activity", []
            else:
                continue                           # assembly / reserved / other
            g = groups.setdefault(subs, {"days": [], "teachers": []})
            g["days"].append(d)
            for t in tch:
                if t not in g["teachers"]:
                    g["teachers"].append(t)
        return [(v["days"], subj, v["teachers"]) for subj, v in groups.items()]

    # ---- render -------------------------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.colors import Color
    from openpyxl.utils import get_column_letter

    def theme_fill(theme, tint):
        return PatternFill(fill_type="solid", fgColor=Color(theme=theme, tint=tint))

    red_fill = PatternFill(fill_type="solid", fgColor=RED_ARGB)
    header_fill = theme_fill(HEADER_THEME, HEADER_TINT)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Master"

    # header row
    for ci, (label, _) in enumerate(plan, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = Font(name=FONT_NAME, bold=True, size=12, color=Color(theme=1))
        cell.fill = red_fill if ci in break_cols else header_fill
        cell.alignment = center
        cell.border = border

    class_ids = sorted(cls_l.keys(), key=lambda c: _class_sort_key(cls_l.get(c, c))) \
        or data.get("classIds", [])

    rows_per_day = 2
    block_height = len(days) * rows_per_day        # 6 days x 2 rows = 12

    for bi, c in enumerate(class_ids):
        top = 2 + bi * block_height
        block_fill = theme_fill(*BLOCK_PALETTE[bi % len(BLOCK_PALETTE)])
        # paint the whole block + borders first (so empty/merged cells carry the colour)
        for r in range(top, top + block_height):
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=r, column=ci)
                cell.fill = red_fill if ci in break_cols else block_fill
                cell.border = border
                cell.font = Font(name=FONT_NAME, size=9)
                cell.alignment = center
            ws.row_dimensions[r].height = 30

        # class label: merge column 1 across the block, centred
        ws.merge_cells(start_row=top, start_column=1, end_row=top + block_height - 1, end_column=1)
        lc = ws.cell(row=top, column=1, value=cls_l.get(c, c))
        lc.font = Font(name=FONT_NAME, bold=True, size=11, color=Color(theme=1))
        lc.alignment = center

        # each period column: subjects grouped across all their days, stacked by earliest
        # day; each is one merged, centred cell sized to its day-count (days x 2 rows).
        for seq, col in seq_col.items():
            r0 = top
            for (dlist, subj, teachers) in day_groups(c, seq):
                span = len(dlist) * rows_per_day
                r1 = r0 + span - 1
                text = f"{subj} ({fmt_days(dlist)})"
                if teachers:
                    text += "\n" + " / ".join(teachers)
                if r1 > r0:
                    ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)
                cell = ws.cell(row=r0, column=col, value=text)
                cell.alignment = center
                cell.font = Font(name=FONT_NAME, size=9)
                r0 = r1 + 1

    # widths + freeze
    ws.column_dimensions["A"].width = 24
    for ci, (label, kind) in enumerate(plan, start=1):
        if kind == "teach":
            ws.column_dimensions[get_column_letter(ci)].width = 36
        elif kind in ("break", "diary"):
            ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "B2"

    wb.save(out)
    title = meta.get("runId", "")
    print(f"wrote {out}{f' (run {title})' if title else ''}: "
          f"{len(class_ids)} classes x {block_height}-row blocks, {ncols} columns (school style)")


if __name__ == "__main__":
    main()
