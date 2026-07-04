#!/usr/bin/env python3
"""
Reverse-engineer a CP-SAT ``solution.json`` from the school's HAND-MADE master
timetable workbook — the inverse of ``render_xls.py``.

Given a run folder (``cpsat/<runId>/`` with ``solver-input.json`` already dumped by the
Node dump worker) and the school's Excel, this reads the Excel grid and emits
``<run-dir>/solution.json`` (a ``Placement[]``) so the existing import -> candidate ->
publish flow can pick it up unchanged (the Python poller treats a folder that already has
``solution.json`` as done, and ``cpsat-import-worker.js`` imports it).

Two phases:
  1. PREFLIGHT (always): resolve every class / subject / teacher / elective band / cohort
     the Excel references against the run's ``solver-input.json``, reconcile weekly period
     counts against the lesson demand, and print an explicit ✓/✗ report + verdict line.
  2. GENERATE: only when the verdict is all-green (or --force) write ``solution.json``.

Design decisions (see the plan / HANDOFF.md):
  - Teacher identity is taken from the run's DB lessons per (class, subject); the Excel
    teacher initials are used only to break ties on split subjects.
  - ``lessonId`` in the output is synthetic ("M1", "M2", ...) — the importer only uses it
    to tie a double together, and we emit singles, so real L# ids are not needed.
  - Doubles are emitted as singles (size 1); the published grid is identical and the demand
    reconciliation counts periods either way. block_group grouping is a display nicety only.

Usage:
    python3 import_excel_solution.py --run-dir cpsat/<runId> \
        --xls "/mnt/h/Time Table 2026-27.xlsx" [--sheet Sheet1] [--dry-run] [--force]

Requires: openpyxl
"""
import argparse
import json
import os
import re
import sys
from collections import defaultdict

# ---------------------------------------------------------------------------
# Small helpers (some mirror render_xls.py)
# ---------------------------------------------------------------------------

DAY_NAMES = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat", 7: "Sun"}
ROMAN_TO_IDX = {"I": 0, "II": 1, "III": 2, "IV": 3, "V": 4, "VI": 5,
                "VII": 6, "VIII": 7, "IX": 8, "X": 9, "XI": 10, "XII": 11}

# Excel-token -> canonical DB subject name(s). Value may be a single string or an ordered
# list of candidates; resolution tries them in order against the class's OWN DB subjects,
# so a token that means different things by grade resolves class-aware:
#   "it"      -> Information Technology (senior) OR Computer (VIII)
#   "art"     -> Art (primary) OR Painting (IX, which has no "Art" subject)
#   "physics" -> Physics (senior) OR Phonics (grade-2 workbook typo "Physics"/"Phonics")
# No class has both variants of any pair, so the ordered class-aware match is unambiguous.
SUBJECT_ALIASES = {
    "it": ["information technology", "computer"],
    "art": ["art", "painting"],
    "physics": ["physics", "phonics"],
    "phe": "physical education",
    "physical edu": "physical education",
    "voc edu": "vocational education",
    "vocedu": "vocational education",
    "v ed": "value education",
    "ved": "value education",
    "evs": "environmental science",
    "gk": "general knowledge",
    "sst": "social studies",
    "dance/music": "music",           # "Music & Dance" — the slash is display, not a band
    "music/dance": "music",
    "maths": "mathematics",
    "cs": "computer science",
    "sulekh": "sulekh",
    "cursive": "cursive",
    "phonics": "phonics",
    "conversation": "conversation",
    "converstion": "conversation",    # workbook typo
    "sanskrit": "sanskrit",
    "reasoning": "reasoning",
    "library": "library",
    "sports": "sports",
    "computer": "computer",
    "hindi": "hindi",
    "science": "science",
    "english": "english",
    "social science": "social science",
    "social studies": "social studies",
    "accountancy": "accountancy",
    "economics": "economics",
    "business studies": "business studies",
    "applied maths": "applied mathematics",
    "applied mathematics": "applied mathematics",
    "chemistry": "chemistry",
    "biology": "biology",
    "painting": "painting",
}

# Tokens that are NOT real subjects (free period / non-teaching markers) — skip.
SKIP_TOKENS = {"activity", "free", "reg", "diary", "break", "lunch", "assembly", "reserved"}

# Excel class-block label -> DB class label (labels.class value). Normalized compare with
# a few explicit aliases for the composite XI streams and section-less single sections.
# keys/values in norm_class() space.
CLASS_ALIASES = {
    "xi sci": "xi-a science",
    "xi com": "xi-a commerce",
    "x": "x-a",
    "ix": "ix-a",
}


def norm(s):
    """lowercase, drop honorifics/codes, collapse punctuation/space for fuzzy compare."""
    s = (s or "").lower()
    s = re.sub(r"\([^)]*\)", " ", s)          # drop "(041-maths)" etc.
    s = s.replace("&", " ").replace(".", " ")
    s = re.sub(r"[^a-z0-9/ ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_class(s):
    """Class-label normalizer that KEEPS the stream so 'XI-A (Science)' and
    'XI-A (Commerce)' stay distinct: -> 'xi-a science' / 'xi-a commerce'.
    Also tightens hyphens so the workbook typo 'VII -B' matches 'VII-B'."""
    s = (s or "").lower().replace("(", " ").replace(")", " ")
    s = re.sub(r"\s*-\s*", "-", s)
    s = re.sub(r"[^a-z0-9\- ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def strip_days(text):
    """'English (2-6)' -> ('English', [2,3,4,5,6]);  'Maths (1, 2)' -> ('Maths',[1,2]).
    Returns (subject_text, [days]) or (None, None) if there's no (day-set) annotation."""
    # first "(day-set)" group; trailing text (e.g. inline teacher initials
    # 'Hindi (1-6) JS') is ignored.
    m = re.search(r"\(\s*([0-9][0-9,\s\-–]*)\)", text)
    if not m:
        return None, None
    subj = text[:m.start()].strip()
    if not subj:
        return None, None
    spec = m.group(1).replace("–", "-")
    days = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        mm = re.match(r"^(\d+)\s*-+\s*(\d+)$", part)   # ranges incl. "1--4" typo
        if mm:
            a, b = int(mm.group(1)), int(mm.group(2))
            days.update(range(min(a, b), max(a, b) + 1))
        elif part.isdigit():
            days.add(int(part))
    return subj, sorted(d for d in days if 1 <= d <= 7)


def name_initials(name):
    """'Mr. Dinesh Chandra Agnihotri' -> 'DCA'."""
    n = re.sub(r"^(mr|mrs|ms|miss)\.?\s+", "", (name or "").strip(), flags=re.I)
    return "".join(w[0] for w in re.split(r"\s+", n) if w).upper()


# ---------------------------------------------------------------------------
# Load run artifacts
# ---------------------------------------------------------------------------

def load_run(run_dir):
    with open(os.path.join(run_dir, "solver-input.json")) as f:
        payload = json.load(f)
    return payload["input"], payload.get("labels", {}), payload.get("meta", {})


def build_grid_maps(inp):
    """Return:
      slot_at[(day, period_idx)] -> (slotId, sequence)   for teaching columns only
      teach_days = sorted day numbers present
    period_idx is 0-based over the teaching columns of each day (I=0..).
    """
    slot_at = {}
    teach_days = []
    for d in inp["grid"]["days"]:
        dow = d["dayOfWeek"]
        teach_days.append(dow)
        teaching = sorted((s["sequence"], s["slotId"]) for s in d["slots"]
                          if s["slotType"] == "teaching")
        for idx, (seq, sid) in enumerate(teaching):
            slot_at[(dow, idx)] = (sid, seq)
    return slot_at, sorted(teach_days)


# ---------------------------------------------------------------------------
# Resolvers built from the run's lessons + labels
# ---------------------------------------------------------------------------

class Resolvers:
    def __init__(self, inp, labels):
        self.inp = inp
        self.labels = labels
        self.class_by_norm = {}          # norm_class(db class label) -> classId
        for cid, lab in labels.get("class", {}).items():
            self.class_by_norm[norm_class(lab)] = cid
        self.subj_name = {sid: lab for sid, lab in labels.get("subject", {}).items()}
        self.teacher_name = {tid: lab for tid, lab in labels.get("teacher", {}).items()}

        # Per class: subjects it teaches (single-class lessons only) -> {norm(name): sid}
        # and teachers per (classId, subjectId).
        self.class_subjects = defaultdict(dict)     # classId -> {normname: sid}
        self.subject_teachers = defaultdict(set)     # (classId, sid) -> {teacherId}
        # Non-band single-class demand: (classId, sid) -> total periods
        self.demand = defaultdict(int)
        # Bands: bandId -> {classIds, offerings(list of {subjectId,teacherId}), demand}
        self.bands = {}
        for L in inp["lessons"]:
            classes = L.get("classIds") or [L["classId"]]
            if L.get("bandId"):
                b = self.bands.setdefault(L["bandId"], {
                    "classIds": classes,
                    "offerings": L["offerings"],
                    "demand": 0,
                })
                b["demand"] += L["size"]
                continue
            for c in classes:
                for o in L["offerings"]:
                    sid = o["subjectId"]
                    self.class_subjects[c][norm(self.subj_name.get(sid, sid))] = sid
                    if o.get("teacherId"):
                        self.subject_teachers[(c, sid)].add(o["teacherId"])
                    self.demand[(c, sid)] += L["size"]

    # -- class -------------------------------------------------------------
    def resolve_class(self, excel_label):
        key = norm_class(excel_label)
        key = CLASS_ALIASES.get(key, key)
        if key in self.class_by_norm:
            return self.class_by_norm[key]
        # section-less single section: "vii" -> "vii-a" if unique
        cands = [cid for k, cid in self.class_by_norm.items() if k == key or k == key + "-a"]
        return cands[0] if len(cands) == 1 else None

    # -- subject (class-aware) ---------------------------------------------
    def _match_in(self, subj_text, cand):
        """Resolve one Excel subject token against a {norm(name): id} candidate map.
        Shared by class-subject resolution (cand = the class's subjects) and elective-band
        matching (cand = a band's offerings). Returns id, "SKIP", or None."""
        want = norm(subj_text)
        if want in SKIP_TOKENS:
            return "SKIP"
        alias = SUBJECT_ALIASES.get(want, want)
        targets = [want] + [t for t in (alias if isinstance(alias, list) else [alias])
                            if t != want]
        # exact match on any target (in priority order)
        for t in targets:
            if t in cand:
                return cand[t]
        # word-set match: token's words must be a SUBSET of the candidate's words. Only
        # this direction — so 'Hindi' <= 'Hindi Course-A' matches, but 'Applied Maths'
        # ({applied,mathematics}) does NOT collapse to 'Mathematics'. Alias-only, so 'IT'
        # never matches 'sanskr-IT'.
        for t in targets:
            tw = set(t.split())
            for cname, cid in cand.items():
                if tw and tw <= set(cname.split()):
                    return cid
        # stream-choice cells like 'PHE/Applied Maths' — take the first part that resolves
        if "/" in subj_text:
            for part in subj_text.split("/"):
                r = self._match_in(part.strip(), cand)
                if r and r != "SKIP":
                    return r
        return None

    def resolve_subject(self, class_id, subj_text):
        return self._match_in(subj_text, self.class_subjects.get(class_id, {}))

    # -- teacher (from lessons, initials break splits) ---------------------
    def resolve_teacher(self, class_id, subject_id, initials_hint):
        ts = self.subject_teachers.get((class_id, subject_id), set())
        if len(ts) == 1:
            return next(iter(ts)), True
        if not ts:
            return None, False
        # split: match the Excel initials
        hint = re.sub(r"[^A-Za-z]", "", (initials_hint or "")).upper()
        matches = [t for t in ts if name_initials(self.teacher_name.get(t, "")) == hint]
        if len(matches) == 1:
            return matches[0], True
        return None, False       # ambiguous / unresolved


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

BLOCK_START = 3        # first block's top row
BLOCK_ROWS = 12        # rows per class block
LABEL_OFFSET = 5       # class label sits at block_start + 5, column 1
NUM_BLOCKS = 20


def cell_text(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def read_period_columns(ws):
    """Row 1 maps period columns: {roman-index -> excel column}. Skips break/Diary."""
    cols = {}
    for c in range(2, ws.max_column + 1):
        head = cell_text(ws, 1, c).upper().replace(".", "")
        if head in ROMAN_TO_IDX:
            cols[ROMAN_TO_IDX[head]] = c
    return cols   # {0:2, 1:3, ... } period_idx -> excel col


def parse_block(ws, top, period_cols):
    """Return list of parsed entries for one class block:
       [{period_idx, days:[...], subject_text, teacher_hint}]
    Read column-by-column; each Subject(day-set) cell pairs with the teacher line below it."""
    entries = []
    rows = list(range(top, top + BLOCK_ROWS))
    for pidx, col in period_cols.items():
        col_cells = [(r, cell_text(ws, r, col)) for r in rows]
        col_cells = [(r, t) for r, t in col_cells if t]
        i = 0
        while i < len(col_cells):
            r, txt = col_cells[i]
            subj, days = strip_days(txt)
            if subj is None:
                i += 1
                continue          # teacher line or stray text; skip
            # teacher hint = the next non-empty cell in this column that is NOT itself a
            # subject(day-set) line.
            hint = ""
            if i + 1 < len(col_cells):
                nr, nt = col_cells[i + 1]
                if strip_days(nt)[0] is None:
                    hint = nt
            entries.append({"period_idx": pidx, "days": days,
                            "subject_text": subj, "teacher_hint": hint})
            i += 1
    return entries


def parse_workbook(ws):
    """Return {excel_class_label: [entries]} for the 20 blocks."""
    period_cols = read_period_columns(ws)
    blocks = {}
    for b in range(NUM_BLOCKS):
        top = BLOCK_START + b * BLOCK_ROWS
        label = cell_text(ws, top + LABEL_OFFSET, 1)
        if not label:
            continue
        blocks[label] = parse_block(ws, top, period_cols)
    return blocks, period_cols


# ---------------------------------------------------------------------------
# Band keyword matching (XI electives only)
# ---------------------------------------------------------------------------

def band_candidates(res):
    """bandId -> {norm(offering subject name): subjectId} for reusing _match_in."""
    out = {}
    for bid, b in res.bands.items():
        out[bid] = {norm(res.subj_name.get(o["subjectId"], "")): o["subjectId"]
                    for o in b["offerings"]}
    return out


def match_band(subj_text, band_cands, res):
    """A slash-cell like 'CS/Painting/Hindi', a stream cell 'Accountancy'/'Biology/Maths',
    or a shared 'English' -> bandId, or None. Resolves each '/'-part against a band's
    offerings with the SAME class-aware matcher used for subjects. A cohort class often
    shows only ITS stream's offering, so one matching non-skip part is enough — but EVERY
    non-skip part must belong to the band (so 'PHE/Applied Maths' never matches a band)."""
    parts = [p.strip() for p in subj_text.split("/") if p.strip()]
    if not parts:
        return None
    for bid, cand in band_cands.items():
        matched_any = False
        ok = True
        for p in parts:
            r = res._match_in(p, cand)
            if r == "SKIP":
                continue
            if r is None:
                ok = False
                break
            matched_any = True
        if ok and matched_any:
            return bid
    return None


def match_band_offerings(subj_text, band_cands, res):
    """Like match_band, but also returns the band offerings the cell actually names
    (a subset of the band's offerings) so we can place ONLY this stream's subjects.
    Science 'Biology/Maths' -> [Biology, Maths]; Commerce 'Accountancy' -> [Accountancy];
    a pooled 'CS/Painting/Hindi' -> all three. Returns (bandId, [offering]) or (None, None)."""
    bid = match_band(subj_text, band_cands, res)
    if not bid:
        return None, None
    cand = band_cands[bid]
    subj_ids = []
    for p in subj_text.split("/"):
        r = res._match_in(p.strip(), cand)
        if r and r != "SKIP" and r not in subj_ids:
            subj_ids.append(r)
    offerings = [o for o in res.bands[bid]["offerings"] if o["subjectId"] in subj_ids]
    return bid, offerings


# idx (0-based teaching period) -> Roman period label, for readable reports.
_ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]


def period_label(pidx):
    return _ROMAN[pidx] if 0 <= pidx < len(_ROMAN) else f"P{pidx + 1}"


# ---------------------------------------------------------------------------
# Build placements + collect preflight facts
# ---------------------------------------------------------------------------

class Report:
    def __init__(self):
        self.classes = []        # (excel_label, classId or None, db_label)
        self.subj_unresolved = []  # (excel_label, token)
        self.teacher_unresolved = []  # (class, subject, hint)
        self.band_rows = []      # (bandId, demand, [(member_label, placed)]) per-stream
        self.cohort_rows = []    # (label, ok)
        self.cohort_align = []   # (cohort_label, aligned_count, [(member_label, day, pidx, band_subjs)])
        self.demand_rows = []    # (class, subject, placed, demand) mismatches only
        self.skipped_slots = []  # (excel_label, day, period_idx, subject) no teaching slot
        self.clashes = []        # (class_label, day, seq, [descr]) two placements same slot
        self.teacher_clashes = []  # (teacher_label, day, plabel, [(class_label, subject)])


def build(res, blocks, slot_at):
    rep = Report()
    band_classes = set()
    for b in res.bands.values():
        band_classes.update(b["classIds"])
    cohort_classes = set()
    for grp in res.inp.get("cohorts", []) or []:
        cohort_classes.update(grp)
    band_cands = band_candidates(res)

    placements = []
    counter = [0]

    def new_id():
        counter[0] += 1
        return "M%d" % counter[0]

    placed_demand = defaultdict(int)     # (classId, sid) -> count (non-band per-class)
    band_placed = defaultdict(int)       # (classId, bandId) -> count (per stream)
    band_slots = defaultdict(lambda: defaultdict(set))  # bandId -> classId -> {(day,pidx)}

    # resolve classes first
    class_of = {}
    for label in blocks:
        cid = res.resolve_class(label)
        rep.classes.append((label, cid, res.labels.get("class", {}).get(cid, "") if cid else ""))
        class_of[label] = cid

    for label, entries in blocks.items():
        cid = class_of[label]
        if not cid:
            continue
        is_band_class = cid in band_classes or cid in cohort_classes
        for e in entries:
            if not e["days"]:
                continue
            # band cell? (cohort classes only). Placed PER STREAM: this class gets the
            # offerings ITS cell names, at ITS own slot — not co-scheduled onto the other
            # stream. Streams coincide where drawn together, diverge where drawn apart.
            bid, boffs = (match_band_offerings(e["subject_text"], band_cands, res)
                          if is_band_class else (None, None))
            for day in e["days"]:
                key = (day, e["period_idx"])
                slot = slot_at.get(key)
                if bid:
                    if slot is None:
                        rep.skipped_slots.append((label, day, e["period_idx"], e["subject_text"]))
                        continue
                    band_placed[(cid, bid)] += 1
                    band_slots[bid][cid].add(key)
                    placements.append({
                        "lessonId": new_id(), "classId": cid, "classIds": None,
                        "dayOfWeek": day, "startSequence": slot[1], "slotIds": [slot[0]],
                        "offerings": boffs, "size": 1, "bandId": bid,
                    })
                    continue
                # per-class subject
                sid = res.resolve_subject(cid, e["subject_text"])
                if sid == "SKIP":
                    continue
                if sid is None:
                    rep.subj_unresolved.append((label, e["subject_text"]))
                    continue
                if slot is None:
                    rep.skipped_slots.append((label, day, e["period_idx"], e["subject_text"]))
                    continue
                tid, ok = res.resolve_teacher(cid, sid, e["teacher_hint"])
                if not ok:
                    rep.teacher_unresolved.append(
                        (label, res.subj_name.get(sid, sid), e["teacher_hint"]))
                placed_demand[(cid, sid)] += 1
                placements.append({
                    "lessonId": new_id(), "classId": cid, "classIds": None,
                    "dayOfWeek": day, "startSequence": slot[1], "slotIds": [slot[0]],
                    "offerings": [{"subjectId": sid, "teacherId": tid}],
                    "size": 1, "bandId": None,
                })

    # -- clash detection (two placements booking the same class at the same slot) ----
    # per-day teaching sequences, to translate a slot's sequence -> period label (I..VIII)
    teach_seq = {d["dayOfWeek"]: sorted(s["sequence"] for s in d["slots"]
                                        if s["slotType"] == "teaching")
                 for d in res.inp["grid"]["days"]}
    seq_of = {s["slotId"]: (day["dayOfWeek"], s["sequence"])
              for day in res.inp["grid"]["days"] for s in day["slots"]}
    occ = defaultdict(list)
    for p in placements:
        for c in (p.get("classIds") or [p["classId"]]):
            for sid in p["slotIds"]:
                occ[(c, sid)].append(p)
    for (c, sid), ps in occ.items():
        if len(ps) > 1:
            day, seq = seq_of.get(sid, (0, 0))
            seqs = teach_seq.get(day, [])
            plabel = period_label(seqs.index(seq)) if seq in seqs else f"seq{seq}"
            descr = [("band " if p["bandId"] else "") +
                     "/".join(res.subj_name.get(o["subjectId"], "").split(" (")[0]
                              for o in p["offerings"]) for p in ps]
            rep.clashes.append((res.labels.get("class", {}).get(c, c), day, plabel, descr))

    # -- teacher clash: one teacher must not be in two classes at the same day+period --
    # Collapse bookings that are the SAME shared band offering (same bandId + subject) to
    # cohort members — that's one physical class (e.g. English taught to both XI streams by
    # one teacher), NOT a clash. Anything else booking a teacher twice at a slot is a clash.
    tocc = defaultdict(list)   # (teacherId, day, slotId) -> [(placement, offering)]
    for p in placements:
        for sid in p["slotIds"]:
            for o in p["offerings"]:
                if o.get("teacherId"):
                    tocc[(o["teacherId"], p["dayOfWeek"], sid)].append((p, o))
    for (tid, day, sid), bookings in tocc.items():
        events = set()          # distinct physical teaching events for this teacher/slot
        for (p, o) in bookings:
            if p["bandId"]:
                events.add(("band", p["bandId"], o["subjectId"]))
            else:
                events.add(("solo", id(p)))     # each standalone class is its own event
        if len(events) > 1:
            seqs = teach_seq.get(day, [])
            _, seq = seq_of.get(sid, (0, 0))
            plabel = period_label(seqs.index(seq)) if seq in seqs else f"seq{seq}"
            who = []
            seen = set()
            for (p, o) in bookings:
                cl = res.labels.get("class", {}).get(p["classId"], p["classId"])
                sb = res.subj_name.get(o["subjectId"], "").split(" (")[0]
                if (cl, sb) not in seen:
                    seen.add((cl, sb))
                    who.append((cl, sb))
            rep.teacher_clashes.append(
                (res.teacher_name.get(tid, tid), day, plabel, who))

    # -- reconcile bands PER STREAM (each member should have periods_per_week) --------
    for bid, b in res.bands.items():
        members = [(res.labels.get("class", {}).get(m, m), band_placed.get((m, bid), 0))
                   for m in b["classIds"]]
        rep.band_rows.append((bid, b["demand"], members))

    # -- cohort alignment: which band slots are shared vs stream-only -----------------
    for grp in res.inp.get("cohorts", []) or []:
        cohort_name = " + ".join(res.labels.get("class", {}).get(c, c) for c in grp)
        aligned = 0
        divergent = []
        for bid in res.bands:
            members = [m for m in grp if band_slots[bid].get(m)]
            if len(members) < 2:
                continue
            all_slots = set().union(*(band_slots[bid][m] for m in members))
            offs = "/".join(res.subj_name.get(o["subjectId"], "").split(" (")[0]
                            for o in res.bands[bid]["offerings"])
            for (day, pidx) in sorted(all_slots):
                have = [m for m in members if (day, pidx) in band_slots[bid][m]]
                if len(have) == len(members):
                    aligned += 1
                else:
                    for m in have:
                        divergent.append(
                            (res.labels.get("class", {}).get(m, m), day, pidx, offs))
        rep.cohort_align.append((cohort_name, aligned, divergent))

    all_keys = set(placed_demand) | set(res.demand)
    for k in sorted(all_keys, key=lambda x: (res.labels.get("class", {}).get(x[0], ""),
                                             res.subj_name.get(x[1], ""))):
        placed, want = placed_demand.get(k, 0), res.demand.get(k, 0)
        if placed != want:
            rep.demand_rows.append((res.labels.get("class", {}).get(k[0], k[0]),
                                    res.subj_name.get(k[1], k[1]), placed, want))

    for grp in res.inp.get("cohorts", []) or []:
        names = [res.labels.get("class", {}).get(c, c) for c in grp]
        rep.cohort_rows.append((" + ".join(names), True))

    return placements, rep


# ---------------------------------------------------------------------------
# Preflight printing + verdict
# ---------------------------------------------------------------------------

def print_report(rep):
    ok_cls = sum(1 for _, cid, _ in rep.classes if cid)
    tot_cls = len(rep.classes)
    print("\n== CLASSES ==")
    for label, cid, db in rep.classes:
        print(f"  {'OK ' if cid else 'XX '} {label:<12} -> {db or '(UNRESOLVED)'}")

    print("\n== ELECTIVE BANDS (per stream) ==")
    for bid, demand, members in rep.band_rows:
        # hard gate = the band was FOUND in the Excel (some stream placed it) with
        # resolvable offerings; a per-stream placed!=demand is a soft review signal.
        found = any(placed > 0 for _, placed in members)
        per = " · ".join(f"{lbl} {placed}/{demand}" for lbl, placed in members)
        diff = any(placed != demand for _, placed in members)
        note = "  (~~ review: a stream's count differs)" if diff else ""
        print(f"  {'OK ' if found else 'XX '} band {bid}: {per}{note}")

    print("\n== COHORTS ==")
    for name, ok in rep.cohort_rows:
        print(f"  {'OK ' if ok else 'XX '} {name}")
    for cohort_name, aligned, divergent in rep.cohort_align:
        print(f"      {cohort_name}: {aligned} band-slot(s) aligned across streams, "
              f"{len(divergent)} stream-only (parallel, different subjects — expected):")
        for lbl, day, pidx, offs in divergent:
            print(f"        · {lbl:<18} {DAY_NAMES.get(day, day)} period {period_label(pidx)} "
                  f"({offs})")

    if rep.subj_unresolved:
        print("\n== UNRESOLVED SUBJECTS ==")
        seen = {}
        for label, tok in rep.subj_unresolved:
            seen[(label, tok)] = seen.get((label, tok), 0) + 1
        for (label, tok), n in sorted(seen.items()):
            print(f"  XX  {label:<12} '{tok}'  (x{n})")

    if rep.teacher_unresolved:
        print("\n== UNRESOLVED / AMBIGUOUS TEACHERS ==")
        seen = set()
        for label, subj, hint in rep.teacher_unresolved:
            if (label, subj, hint) in seen:
                continue
            seen.add((label, subj, hint))
            print(f"  XX  {label:<12} {subj:<28} hint='{hint}'")

    if rep.demand_rows:
        print("\n== DEMAND RECONCILIATION (review — not a blocker) ==")
        print("   placed = periods read from the Excel; demand = periods in the solver's")
        print("   lessons. Differences are usually a genuine hand-vs-solver difference or a")
        print("   source cell missing its (day-set). Review, then fix the Excel if wrong.")
        for cls, subj, placed, want in rep.demand_rows:
            print(f"  ~~  {cls:<18} {subj:<30} placed {placed} / demand {want}")

    if rep.clashes:
        print("\n== CLASS/SLOT CLASHES (review — two subjects booked at once) ==")
        print("   Usually the two cohort streams disagree on an elective-band slot, or a")
        print("   subject overlaps a band. Fix the Excel so the streams align.")
        for cls, day, plabel, descr in rep.clashes:
            print(f"  !!  {cls:<18} {DAY_NAMES.get(day, day)} period {plabel}: "
                  f"{'  vs  '.join(descr)}")

    if rep.teacher_clashes:
        print("\n== TEACHER CLASHES (review — one teacher, two classes at once) ==")
        print("   A teacher is booked in more than one class at the same day+period.")
        print("   Fix the Excel (move one of the subjects to a free period).")
        for tname, day, plabel, who in rep.teacher_clashes:
            pairs = "  vs  ".join(f"{cl} {sb}" for cl, sb in who)
            print(f"  !!  {tname:<24} {DAY_NAMES.get(day, day)} period {plabel}: {pairs}")

    if rep.skipped_slots:
        print("\n== CELLS WITH NO TEACHING SLOT (skipped) ==")
        for label, day, pidx, subj in rep.skipped_slots[:20]:
            print(f"  --  {label:<12} day{day} period#{pidx+1} '{subj}'")
        if len(rep.skipped_slots) > 20:
            print(f"      ... and {len(rep.skipped_slots) - 20} more")

    bands_found = sum(1 for _, _, members in rep.band_rows
                      if any(placed > 0 for _, placed in members))
    bands_all_found = bands_found == len(rep.band_rows)
    band_count_diffs = sum(1 for _, demand, members in rep.band_rows
                           if any(placed != demand for _, placed in members))
    # Hard gate for writing = every ENTITY the Excel references resolves (classes,
    # subjects, teachers, and each band recognized). Count differences (demand + band
    # alignment) are soft review — the hand timetable may legitimately differ.
    green = (ok_cls == tot_cls and not rep.subj_unresolved and not rep.teacher_unresolved
             and bands_all_found)
    review = len(rep.demand_rows) + band_count_diffs
    print("\n" + "=" * 70)
    print(f"RESOLVE: classes {ok_cls}/{tot_cls} {'OK' if ok_cls == tot_cls else 'XX'} | "
          f"bands {bands_found}/{len(rep.band_rows)} {'OK' if bands_all_found else 'XX'} | "
          f"cohorts {sum(1 for _, o in rep.cohort_rows if o)}/{len(rep.cohort_rows)} | "
          f"subjects {'OK' if not rep.subj_unresolved else 'XX'} | "
          f"teachers {'OK' if not rep.teacher_unresolved else 'XX'}")
    if review:
        print(f"REVIEW:  {review} count difference(s) vs solver lessons "
              f"({len(rep.demand_rows)} demand, {band_count_diffs} band) — not blocking.")
    if rep.clashes:
        print(f"CLASHES: {len(rep.clashes)} class/slot clash(es) — review before publishing.")
    if rep.teacher_clashes:
        print(f"TEACHER: {len(rep.teacher_clashes)} teacher double-booking(s) — review before publishing.")
    print(f"VERDICT: {'ALL ENTITIES RESOLVED' if green else 'NOT READY — fix the XX items above'}")
    print("=" * 70)
    return green


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", required=True, help="cpsat/<runId> (has solver-input.json)")
    ap.add_argument("--xls", required=True, help="path to the hand-made timetable workbook")
    ap.add_argument("--sheet", default="Sheet1")
    ap.add_argument("--out", default=None, help="output solution.json (default <run-dir>/solution.json)")
    ap.add_argument("--dry-run", action="store_true", help="preflight only; write nothing")
    ap.add_argument("--force", action="store_true", help="write even if preflight is not all-green")
    ap.add_argument("--debug-class", default=None, help="dump raw parsed entries for one class block and exit")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(args.xls, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active

    inp, labels, meta = load_run(args.run_dir)
    res = Resolvers(inp, labels)
    slot_at, _ = build_grid_maps(inp)

    blocks, period_cols = parse_workbook(ws)
    if args.debug_class:
        band_cands = band_candidates(res)
        for label, entries in blocks.items():
            if norm_class(label) != norm_class(args.debug_class):
                continue
            cid = res.resolve_class(label)
            print(f"-- {label} -> {res.labels.get('class', {}).get(cid, cid)} "
                  f"(period_cols={period_cols}) --")
            tally = defaultdict(int)
            for e in sorted(entries, key=lambda x: (x["period_idx"], x["days"])):
                bid = match_band(e["subject_text"], band_cands, res)
                if bid:
                    dest = "BAND " + bid
                else:
                    sid = res.resolve_subject(cid, e["subject_text"]) if cid else None
                    dest = "SKIP" if sid == "SKIP" else (
                        res.subj_name.get(sid, sid) if sid else "??UNRESOLVED")
                    if sid and sid != "SKIP":
                        tally[dest] += len(e["days"])
                print(f"  P{e['period_idx']+1} days={e['days']} "
                      f"subj='{e['subject_text']}' -> {dest}")
            print("  TALLY:", dict(sorted(tally.items())))
        return 0
    placements, rep = build(res, blocks, slot_at)
    green = print_report(rep)
    print(f"\nplacements built: {len(placements)}")

    # Writable unless an ENTITY is unresolved (green) or the timetable has a clash (a
    # clashing grid is invalid to publish). Count/demand differences do NOT block.
    blockers = []
    if not green:
        blockers.append("unresolved entities")
    if rep.clashes:
        blockers.append(f"{len(rep.clashes)} class/slot clash(es)")
    if rep.teacher_clashes:
        blockers.append(f"{len(rep.teacher_clashes)} teacher double-booking(s)")

    if args.dry_run:
        print("(--dry-run: nothing written)")
        return 0 if not blockers else 1
    if blockers and not args.force:
        print(f"Refusing to write solution.json ({', '.join(blockers)}). "
              f"Fix the Excel, or pass --force to override.")
        return 1

    out = args.out or os.path.join(args.run_dir, "solution.json")
    with open(out, "w") as f:
        json.dump(placements, f, indent=2)
    print(f"wrote {out}: {len(placements)} placements")
    return 0


if __name__ == "__main__":
    sys.exit(main())

